"""
create_template.py
Generates two Excel files users can download from the app:
  1. hr_template_blank.xlsx  — empty sheet with all supported columns + instructions
  2. hr_template_sample.xlsx — 50 pre-filled example rows to show expected format
"""

import datetime as dt
import io
import numpy as np
import pandas as pd

# ── All supported columns with explanations ────────────────────────────────────
COLUMNS = [
    # Core identity
    ("EmployeeID",          "Unique employee identifier (e.g. EMP001)",                      "EMP001"),
    ("EmployeeName",        "Full name (optional — not used in analysis)",                   "John Smith"),
    # Org hierarchy
    ("Division",            "Top-level business division (e.g. Pharmaceuticals, Logistics)", "Pharmaceuticals"),
    ("BusinessUnit",        "Sub-division or business unit within the division",              "Pharma Sales"),
    ("DepartmentGroup",     "Department grouping (optional, for large orgs)",                "Sales — North"),
    ("Department",          "Department name (e.g. Sales, Engineering, HR)",                 "Retail Sales"),
    ("SubDepartment",       "Sub-department (optional)",                                     ""),
    # Role
    ("DesignationLevel",    "Management level: Executive/Officer | Manager | Director etc.", "Manager"),
    ("Designation",         "Job title / designation (e.g. Senior Sales Executive)",         "Senior Sales Executive"),
    ("Grade",               "Pay grade or band code (e.g. M3, G4, Band-B)",                 "M3"),
    # Employment details
    ("EmploymentType",      "Permanent | Probationary | Contractual | Graded",               "Permanent"),
    ("WorkLocation",        "Office | Field | Factory | Remote | Head Office",               "Field"),
    ("Location",            "City or office name (e.g. Sydney, Dhaka)",                      "Sydney"),
    ("District",            "District or region of the employee's home",                     "Dhaka"),
    # Demographics
    ("Gender",              "Male | Female | Non-binary",                                    "Male"),
    ("Religion",            "Religion (optional — used for diversity analysis)",             "Islam"),
    ("MaritalStatus",       "Married | Single | Divorced | Widowed",                         "Married"),
    ("Degree",              "Highest qualification: BSc | MBA | HSC | PhD etc.",             "MBA"),
    ("Subject",             "Field of study (e.g. Finance, Engineering)",                    "Marketing"),
    # Dates
    ("BirthDate",           "Date of birth — YYYY-MM-DD format",                             "1990-05-14"),
    ("JoiningDate",         "Date the employee joined — YYYY-MM-DD",                         "2019-03-01"),
    ("LeftDate",            "Date the employee left (blank if still active) — YYYY-MM-DD",   ""),
    ("ConfirmDate",         "Probation confirmation date — YYYY-MM-DD (blank if unconfirmed)","2019-09-01"),
    ("LstPromotionDate",    "Date of last promotion — YYYY-MM-DD (blank if never promoted)", "2022-06-15"),
    # Status
    ("Active",              "Y = currently employed | N = has left",                          "Y"),
    # Manager
    ("SuperCode",           "Employee ID of the direct manager (e.g. EMP007)",               "EMP007"),
    # Recruitment
    ("RecruitmentSource",   "LinkedIn | Referral | Job Board | Recruiter Agency | University","LinkedIn"),
    ("TimeToHireDays",      "Number of days from job posting to hiring (integer)",            "28"),
    # Compensation
    ("Salary",              "Annual base salary (number only, no currency symbol)",           "85000"),
    ("SalaryRangeMax",      "Maximum of the salary band for this role",                      "100000"),
    ("Bonus",               "Annual bonus amount (number only)",                              "8500"),
    # Performance
    ("PerformanceRating",   "Overall performance rating — integer 1 (low) to 5 (high)",      "4"),
    ("EngagementScore",     "Employee engagement score — 0 to 100",                          "72"),
    ("ApprisalPoint2019",   "Appraisal score for 2019 (0–100). Leave blank if not applicable","68"),
    ("ApprisalPoint2020",   "Appraisal score for 2020 (0–100)",                              "74"),
    ("ApprisalPoint2021",   "Appraisal score for 2021 (0–100)",                              "79"),
    # Leave
    ("PreLeaveEntitled",    "Annual privileged leave entitlement (days)",                    "20"),
    ("PreLeaveDAYSTAKEN",   "Privileged leave days actually taken this year",                "12"),
    ("PreLeaveBalance",     "Remaining privileged leave balance",                            "8"),
    ("SECLeaveDAYSTAKEN",   "Sick / casual leave days taken this year",                      "5"),
    ("SecLeaveBalance",     "Remaining sick / casual leave balance",                         "9"),
    # Workload
    ("OvertimeHoursMonth",  "Average overtime hours worked per month",                       "8.5"),
    ("TrainingHoursYear",   "Total training hours completed this year",                      "24"),
    ("AbsenceDaysYear",     "Total unplanned absence days this year",                        "6"),
    ("DistanceFromOfficeKm","Distance from home to office (km)",                             "12"),
]

COLUMN_NAMES = [c[0] for c in COLUMNS]
DESCRIPTIONS = [c[1] for c in COLUMNS]
EXAMPLES     = [c[2] for c in COLUMNS]


def _style_wb(wb):
    """Apply professional styling to the workbook."""
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, numbers as xl_num)
    from openpyxl.utils import get_column_letter

    ACCENT   = "1A1F36"   # Dark navy
    ACCENT2  = "5046E4"   # Indigo
    LIGHT    = "EEF2FF"   # Light indigo
    GREY_ROW = "F8FAFC"
    BORDER_C = "E2E8F0"

    thin   = Side(style="thin",   color=BORDER_C)
    thick  = Side(style="medium", color=ACCENT)
    border = Border(bottom=thin)

    for ws in wb.worksheets:
        # Header row
        for cell in ws[1]:
            cell.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
            cell.fill      = PatternFill("solid", fgColor=ACCENT)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border    = Border(bottom=Side(style="medium", color=ACCENT2))

        # Data rows
        for ri, row in enumerate(ws.iter_rows(min_row=2), start=2):
            is_alt = ri % 2 == 0
            for cell in row:
                cell.font      = Font(name="Calibri", size=10, color="111827")
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border    = border
                if is_alt:
                    cell.fill = PatternFill("solid", fgColor=GREY_ROW)

        # Column widths
        for col in ws.columns:
            max_w = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_w + 4, 14), 42)

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28

    return wb


def build_blank_template() -> bytes:
    """
    Empty template with column headers, descriptions, and example row.
    Users delete the example row and fill in their data.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import openpyxl.styles.numbers

    wb = Workbook()

    # ── Sheet 1: Data Entry ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Employee Data"

    # Row 1: Column headers
    for col_idx, name in enumerate(COLUMN_NAMES, 1):
        ws.cell(row=1, column=col_idx, value=name)

    # Row 2: Example row (pre-filled, styled differently so user knows to replace)
    for col_idx, example in enumerate(EXAMPLES, 1):
        ws.cell(row=2, column=col_idx, value=example)

    # Rows 3–52: Blank data entry rows (50 rows ready to fill)
    # (just empty, styling applied by _style_wb)

    # ── Sheet 2: Column Guide ──────────────────────────────────────────────────
    ws2 = wb.create_sheet("Column Guide")
    ws2.append(["Column Name", "Description", "Example Value", "Required?"])
    required_cols = {"EmployeeID","JoiningDate","Active","Department","Gender","Salary","PerformanceRating","EngagementScore"}
    for name, desc, example in COLUMNS:
        req = "✅ Required" if name in required_cols else "Optional"
        ws2.append([name, desc, example, req])

    # ── Sheet 3: Valid Values ──────────────────────────────────────────────────
    ws3 = wb.create_sheet("Valid Values")
    ws3.append(["Field", "Accepted Values"])
    valid_vals = [
        ("Active",            "Y | N"),
        ("Gender",            "Male | Female | Non-binary"),
        ("EmploymentType",    "Permanent | Probationary | Contractual | Graded | Casual"),
        ("WorkLocation",      "Office | Field | Factory | Remote | Head Office"),
        ("MaritalStatus",     "Married | Single | Divorced | Widowed"),
        ("DesignationLevel",  "Executive / Officer | Assistant Manager | Manager | General Manager | Director"),
        ("RecruitmentSource", "LinkedIn | Referral | Job Board | Recruiter Agency | University | Company Website"),
        ("PerformanceRating", "1 | 2 | 3 | 4 | 5  (1=lowest, 5=highest)"),
        ("EngagementScore",   "0 – 100"),
        ("Dates",             "YYYY-MM-DD format (e.g. 2023-06-15). Leave blank if not applicable."),
        ("Salary / Pay",      "Numbers only — no currency symbol. Use the currency of your country."),
    ]
    for row in valid_vals:
        ws3.append(row)

    # ── Sheet 4: Instructions ──────────────────────────────────────────────────
    ws4 = wb.create_sheet("How To Use")
    instructions = [
        ("GETTING STARTED", ""),
        ("Step 1", "Fill in your employee data in the 'Employee Data' sheet, starting from Row 3."),
        ("Step 2", "Delete Row 2 (the example row) when you have added your real data."),
        ("Step 3", "Keep column headers in Row 1 exactly as they are — do not rename or remove them."),
        ("Step 4", "Save the file and upload it to the HR Analytics Platform."),
        ("", ""),
        ("TIPS", ""),
        ("Minimum data needed",    "EmployeeID, JoiningDate, Active, and at least one grouping column (e.g. Department)."),
        ("More columns = more charts", "The more columns you fill in, the more analytics will be unlocked automatically."),
        ("Dates",                   "Always use YYYY-MM-DD format (e.g. 2023-06-15). Do not use formatted dates like '15/06/2023'."),
        ("Currency",                "Enter salary as a plain number (e.g. 85000). Do not include $ or any currency symbol."),
        ("Missing data",            "Leave a cell blank if you don't have that information. Do not enter 0 or N/A."),
        ("Case sensitivity",        "Gender, Active, EmploymentType etc. are not case-sensitive — Male/male/MALE all work."),
        ("Large datasets",          "The app supports files up to 200MB. For best performance, keep to under 100,000 rows."),
        ("Multi-division orgs",     "Fill in Division, BusinessUnit, DepartmentGroup, Department to enable the full hierarchy drill-down."),
        ("", ""),
        ("COLUMN REQUIREMENT LEVELS", ""),
        ("Required",                "EmployeeID, JoiningDate, Active, Department, Gender, Salary, PerformanceRating, EngagementScore"),
        ("Strongly recommended",    "Division, BusinessUnit, DesignationLevel, EmploymentType, BirthDate, LeftDate"),
        ("Optional",                "All other columns — fill in as many as possible for richer analytics"),
    ]
    ws4.append(["Topic", "Guidance"])
    for row in instructions:
        ws4.append(row)

    wb = _style_wb(wb)

    # Highlight example row in yellow so user knows to delete/replace it
    from openpyxl.styles import Font, PatternFill
    for col_idx in range(1, len(COLUMN_NAMES)+1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = PatternFill("solid", fgColor="FEF9C3")
        cell.font = Font(name="Calibri", size=10, color="92400E", italic=True)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


def build_sample_50() -> bytes:
    """50-row pre-filled sample showing the expected data format."""
    from sample_data import generate_sample_data
    import pandas as pd

    df = generate_sample_data(n=50, seed=99)

    # Rename columns to match template
    rename_map = {
        "EmployeeID":"EmployeeID","EmployeeName":"EmployeeName",
        "Division":"Division","DeptUnit":"BusinessUnit",
        "DepartmentGroup":"DepartmentGroup","DeptName":"Department",
        "DesgMainName":"DesignationLevel","DesgName":"Designation",
        "Grade":"Grade","JobTypeDetails":"EmploymentType",
        "WorkLocation":"WorkLocation","Location":"Location",
        "DistrictName":"District","Gender":"Gender",
        "Religion":"Religion","MaritalStatus":"MaritalStatus",
        "Degree":"Degree","BirthDate":"BirthDate",
        "JoiningDate":"JoiningDate","LeftDate":"LeftDate",
        "ConfirmDate":"ConfirmDate","LstPromotionDate":"LstPromotionDate",
        "Active":"Active","SuperCode":"SuperCode",
        "RecruitmentSource":"RecruitmentSource","TimeToHireDays":"TimeToHireDays",
        "Salary":"Salary","SalaryRangeMax":"SalaryRangeMax","Bonus":"Bonus",
        "PerformanceRating":"PerformanceRating","EngagementScore":"EngagementScore",
        "ApprisalPoint2019":"ApprisalPoint2019","ApprisalPoint2020":"ApprisalPoint2020",
        "ApprisalPoint2021":"ApprisalPoint2021",
        "PreLeaveEntitled":"PreLeaveEntitled","PreLeaveDAYSTAKEN":"PreLeaveDAYSTAKEN",
        "PreLeaveBalance":"PreLeaveBalance","SECLeaveDAYSTAKEN":"SECLeaveDAYSTAKEN",
        "SecLeaveBalance":"SecLeaveBalance",
        "OvertimeHoursMonth":"OvertimeHoursMonth","TrainingHoursYear":"TrainingHoursYear",
        "AbsenceDaysYear":"AbsenceDaysYear","DistanceFromOfficeKm":"DistanceFromOfficeKm",
        "Attrition":"Active_Flag",  # we'll derive Active from Attrition
    }
    df = df.rename(columns={k:v for k,v in rename_map.items() if k in df.columns})
    df["Active"] = df["Attrition"].map({"Yes":"N","No":"Y"}) if "Attrition" in df.columns else "Y"

    # Keep only template columns that exist
    out_cols = [c for c in COLUMN_NAMES if c in df.columns]
    df_out = df[out_cols].copy()

    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = "Employee Data"
    # Headers
    for ci, name in enumerate(out_cols, 1):
        ws.cell(row=1, column=ci, value=name)
    # Data
    for ri, row in enumerate(df_out.itertuples(index=False), start=2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value="" if pd.isna(val) else val)

    # Add guide sheet
    ws2 = wb.create_sheet("Column Guide")
    ws2.append(["Column Name", "Description", "Example"])
    for name, desc, example in COLUMNS:
        ws2.append([name, desc, example])

    wb = _style_wb(wb)
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()


if __name__ == "__main__":
    blank = build_blank_template()
    with open("hr_template_blank.xlsx","wb") as f: f.write(blank)
    sample = build_sample_50()
    with open("hr_template_sample.xlsx","wb") as f: f.write(sample)
    print(f"✅ hr_template_blank.xlsx  — {len(blank):,} bytes")
    print(f"✅ hr_template_sample.xlsx — {len(sample):,} bytes")
